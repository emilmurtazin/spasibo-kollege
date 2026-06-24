import calendar
from datetime import date

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login as auth_login, logout as auth_logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction as db_transaction
from django.db.models import Sum, Count, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.urls import reverse
from django.utils import timezone

from .forms import (
    CompanyRegistrationForm, EmailAuthenticationForm, GiveCoinsForm,
    RewardForm, GoalForm, ENPSResponseForm,
)
from .models import Company, User, Reward, Transaction, ENPSSurvey, CompanyInvite, TelegramLinkToken, SeniorityBonus


def admin_required(view_func):
    """Декоратор: доступ только для администраторов компании (роль admin)."""
    return user_passes_test(lambda u: u.is_authenticated and u.role == 'admin', login_url='dashboard')(view_func)


# ---------------------------------------------------------------------------
# Публичные страницы
# ---------------------------------------------------------------------------

def landing(request):
    """Лендинг с описанием платформы и тарифами."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'core/landing.html', {
        'plans': settings.SUBSCRIPTION_PLANS,
    })


def register_company(request):
    """Регистрация новой компании и первого администратора."""
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        form = CompanyRegistrationForm(request.POST)
        if not request.POST.get('consent'):
            form.add_error(None, 'Необходимо принять пользовательское соглашение и согласие на обработку данных.')
        if form.is_valid():
            data = form.cleaned_data
            plan_info = settings.SUBSCRIPTION_PLANS[data['subscription_plan']]

            company = Company.objects.create(
                name=data['company_name'],
                inn=data.get('inn', ''),
                subscription_plan=data['subscription_plan'],
                employee_limit=plan_info['employee_limit'],
            )

            # Дефолтные правила бонусов за стаж — администратор сможет изменить позже
            SeniorityBonus.objects.bulk_create([
                SeniorityBonus(company=company, days_required=90,  coins_amount=15),
                SeniorityBonus(company=company, days_required=180, coins_amount=30),
                SeniorityBonus(company=company, days_required=365, coins_amount=50),
                SeniorityBonus(company=company, days_required=730, coins_amount=75),
                SeniorityBonus(company=company, days_required=1095, coins_amount=100),
            ])

            user = User.objects.create_user(
                username=data['email'],
                email=data['email'],
                password=data['password1'],
                first_name=data['first_name'],
                last_name=data['last_name'],
                company=company,
                role='admin',
                hire_date=timezone.now().date(),
                balance=settings.MONTHLY_COIN_ALLOCATION,
                last_monthly_allocation=timezone.now().date(),
            )

            auth_login(request, user)
            messages.success(request, f'Компания успешно зарегистрирована!')
            return redirect('onboarding')
    else:
        form = CompanyRegistrationForm()

    return render(request, 'core/register.html', {'form': form})


def login_view(request):
    """Вход по email и паролю."""
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        form = EmailAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            email = form.cleaned_data['username']
            password = form.cleaned_data['password']
            try:
                user_obj = User.objects.get(email=email)
            except User.DoesNotExist:
                user_obj = None

            user = None
            if user_obj is not None:
                user = authenticate(request, username=user_obj.username, password=password)

            if user is not None:
                auth_login(request, user)
                return redirect('dashboard')
            messages.error(request, 'Неверный email или пароль.')
    else:
        form = EmailAuthenticationForm()

    return render(request, 'core/login.html', {'form': form})


def logout_view(request):
    auth_logout(request)
    return redirect('landing')


# ---------------------------------------------------------------------------
# Сотрудник: дашборд, отправка спасибо, лента, топ, награды, профиль
# ---------------------------------------------------------------------------

@login_required
def dashboard(request):
    """Личный дашборд сотрудника."""
    user = request.user
    today = timezone.now().date()

    recent_received = Transaction.objects.filter(
        to_user=user, type='give'
    ).select_related('from_user')[:5]

    recent_sent = Transaction.objects.filter(
        from_user=user, type='give'
    ).select_related('to_user')[:5]

    coins_received_total = Transaction.objects.filter(
        to_user=user, type='give'
    ).aggregate(total=Sum('amount'))['total'] or 0

    coins_sent_total = Transaction.objects.filter(
        from_user=user, type='give'
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Монетки от коллег (type=give) — идут в зачёт цели
    coins_from_peers = Transaction.objects.filter(
        to_user=user, type='give'
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Бонусы за стаж (type=bonus) — идут в зачёт цели
    coins_from_bonus = Transaction.objects.filter(
        to_user=user, type='bonus'
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Ежемесячные системные начисления (type=admin) — НЕ в зачёт цели
    coins_from_system = Transaction.objects.filter(
        to_user=user, type='admin'
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Потрачено на награды
    coins_spent = Transaction.objects.filter(
        from_user=user, type='reward'
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Монетки в зачёт цели = от коллег + бонусы за стаж − потраченные
    coins_toward_goal = max(0, coins_from_peers + coins_from_bonus - coins_spent)

    progress_percent = None
    if user.target_reward and user.target_reward.price:
        progress_percent = min(100, int(coins_toward_goal / user.target_reward.price * 100))

    # Активный опрос eNPS, на который сотрудник ещё не ответил
    from .models import ENPSParticipation
    pending_survey = None
    if user.company:
        latest_survey = ENPSSurvey.objects.filter(company=user.company).order_by('-started_at').first()
        if latest_survey and not ENPSParticipation.objects.filter(survey=latest_survey, user=user).exists():
            pending_survey = latest_survey

    # Сплочённость отделов: своего — отдельно для акцента, плюс полный список всех отделов
    my_department_cohesion = None
    all_department_cohesion = []
    if user.company:
        all_department_cohesion = _calculate_department_cohesion(user.company, today.year, today.month)
        if user.department:
            my_department_cohesion = next(
                (d for d in all_department_cohesion if d['department'] == user.department), None
            )
        # Остальные отделы (кроме своего) — для общего списка ниже
        other_department_cohesion = [
            d for d in all_department_cohesion if d['department'] != user.department
        ]
    else:
        other_department_cohesion = []

    context = {
        'user_obj': user,
        'pending_survey': pending_survey,
        'my_department_cohesion': my_department_cohesion,
        'other_department_cohesion': other_department_cohesion,
        'recent_received': recent_received,
        'recent_sent': recent_sent,
        'coins_received_total': coins_received_total,
        'coins_sent_total': coins_sent_total,
        'coins_from_peers': coins_from_peers,
        'coins_from_bonus': coins_from_bonus,
        'coins_from_system': coins_from_system,
        'coins_toward_goal': coins_toward_goal,
        'progress_percent': progress_percent,
        'today': today,
    }
    return render(request, 'core/dashboard.html', context)


@login_required
def give_coins(request):
    """Страница отправки «Спасибо» коллеге."""
    user = request.user
    today = timezone.now().date()

    if request.method == 'POST':
        form = GiveCoinsForm(request.POST, sender=user)
        if form.is_valid():
            to_user = form.cleaned_data['to_user']
            amount = form.cleaned_data['amount']
            comment = form.cleaned_data['comment']

            # Проверка: достаточно ли монет у отправителя
            if amount > user.balance:
                form.add_error('amount', 'Недостаточно монет на балансе.')
            else:
                # Проверка: лимит 15 монет одному коллеге в месяц
                already_given = user.coins_given_to(to_user, today.year, today.month)
                limit = settings.MAX_COINS_PER_RECEIVER_PER_MONTH
                if already_given + amount > limit:
                    remaining = max(0, limit - already_given)
                    form.add_error(
                        'amount',
                        f'Лимит {limit} монет одному коллеге в месяц. '
                        f'Вы уже отправили {already_given}, можно ещё максимум {remaining}.'
                    )
                else:
                    with db_transaction.atomic():
                        user.balance -= amount
                        user.save(update_fields=['balance'])

                        to_user.balance += amount
                        to_user.save(update_fields=['balance'])

                        Transaction.objects.create(
                            from_user=user,
                            to_user=to_user,
                            amount=amount,
                            comment=comment,
                            type='give',
                            month=today.month,
                            year=today.year,
                        )
                    # Telegram-уведомление получателю
                    try:
                        from .telegram_bot import notify_coins_received, notify_reward_available
                        notify_coins_received(to_user, user, amount, comment)
                        notify_reward_available(to_user)
                    except Exception:
                        pass
                    messages.success(request, f'Спасибо отправлено! {to_user} получил {amount} монет.')
                    return redirect('give_coins')
    else:
        form = GiveCoinsForm(sender=user)

    # Сколько уже отправлено каждому коллеге в этом месяце (для подсказки)
    sent_this_month = Transaction.objects.filter(
        from_user=user, type='give', year=today.year, month=today.month
    ).values('to_user').annotate(total=Sum('amount'))
    sent_map = {row['to_user']: row['total'] for row in sent_this_month}

    return render(request, 'core/give_coins.html', {
        'form': form,
        'limit': settings.MAX_COINS_PER_RECEIVER_PER_MONTH,
        'sent_map': sent_map,
    })


@login_required
def feed(request):
    """Лента признаний по компании."""
    qs = Transaction.objects.filter(
        type='give', from_user__company=request.user.company
    ).select_related('from_user', 'to_user')

    department = request.GET.get('department', '')
    if department:
        qs = qs.filter(
            Q(from_user__department=department) | Q(to_user__department=department)
        )

    departments = User.objects.filter(
        company=request.user.company
    ).exclude(department='').values_list('department', flat=True).distinct()

    return render(request, 'core/feed.html', {
        'transactions': qs[:100],
        'departments': departments,
        'selected_department': department,
    })


@login_required
def top_employees(request):
    """Топ сотрудников по полученным и отправленным монетам за период."""
    period = request.GET.get('period', 'month')
    today = timezone.now().date()

    qs = Transaction.objects.filter(type='give', from_user__company=request.user.company)

    if period == 'month':
        qs = qs.filter(year=today.year, month=today.month)
        period_label = 'за этот месяц'
    elif period == 'year':
        qs = qs.filter(year=today.year)
        period_label = 'за этот год'
    else:
        period_label = 'за всё время'

    top_receivers = qs.values(
        'to_user__id', 'to_user__first_name', 'to_user__last_name', 'to_user__department'
    ).annotate(total=Sum('amount')).order_by('-total')[:10]

    top_senders = qs.values(
        'from_user__id', 'from_user__first_name', 'from_user__last_name', 'from_user__department'
    ).annotate(total=Sum('amount')).order_by('-total')[:10]

    return render(request, 'core/top_employees.html', {
        'top_receivers': top_receivers,
        'top_senders': top_senders,
        'period': period,
        'period_label': period_label,
    })


@login_required
def rewards_shop(request):
    """Магазин наград — список доступных наград и покупка за монеты.

    Покупка возможна только за монетки, полученные от коллег (type=give)
    или в виде бонуса за стаж (type=bonus). Ежемесячные системные начисления
    (type=admin) в зачёт покупки не идут.
    """
    rewards = Reward.objects.filter(company=request.user.company, is_active=True)
    user    = request.user

    purchasable_balance = user.purchasable_balance()

    if request.method == 'POST':
        reward = get_object_or_404(Reward, pk=request.POST.get('reward_id'), company=request.user.company)

        if reward.price > purchasable_balance:
            messages.error(
                request,
                f'Недостаточно монет от коллег для «{reward.name}». '
                f'Нужно {reward.price}, доступно для покупки: {purchasable_balance}. '
                f'Ежемесячные начисления нельзя тратить на награды — только монетки от коллег и бонусы за стаж.'
            )
        elif reward.price > user.balance:
            messages.error(request, f'Недостаточно монет для «{reward.name}». Нужно {reward.price}, у вас {user.balance}.')
        else:
            with db_transaction.atomic():
                user.balance -= reward.price
                user.save(update_fields=['balance'])
                today = timezone.now().date()
                Transaction.objects.create(
                    from_user=user,
                    to_user=None,
                    reward=reward,
                    status='pending',
                    amount=reward.price,
                    comment=f'Покупка награды: {reward.name}',
                    type='reward',
                    month=today.month,
                    year=today.year,
                )
            messages.success(
                request,
                f'Награда «{reward.name}» заказана! Администратор получит уведомление '
                f'и подтвердит выдачу — статус можно отследить в профиле.'
            )

            # Уведомление HR-администраторам компании
            try:
                from .telegram_bot import notify_reward_purchased
                admins = User.objects.filter(company=user.company, role='admin', is_active=True)
                for admin_user in admins:
                    if admin_user.telegram_chat_id:
                        notify_reward_purchased(admin_user, user, reward)
            except Exception:
                pass
        return redirect('rewards_shop')

    # История покупок текущего сотрудника (для отображения статуса)
    my_purchases = Transaction.objects.filter(
        from_user=user, type='reward'
    ).select_related('reward').order_by('-date')[:10]

    return render(request, 'core/rewards_shop.html', {
        'rewards': rewards,
        'purchasable_balance': purchasable_balance,
        'my_purchases': my_purchases,
    })


@login_required
def profile(request):
    """Профиль сотрудника: выбор цели + самостоятельное редактирование данных."""
    from .forms import ProfileEditForm
    from django.contrib.auth import update_session_auth_hash

    user = request.user
    edit_form = ProfileEditForm(instance=user)
    goal_form = GoalForm(instance=user, company=user.company)

    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'edit_profile':
            edit_form = ProfileEditForm(request.POST, instance=user)
            if edit_form.is_valid():
                password_changed = bool(edit_form.cleaned_data.get('new_password'))
                edit_form.save()
                if password_changed:
                    update_session_auth_hash(request, user)  # не разлогинивать после смены пароля
                messages.success(request, 'Данные профиля обновлены.')
                return redirect('profile')
        else:
            goal_form = GoalForm(request.POST, instance=user, company=user.company)
            if goal_form.is_valid():
                goal_form.save()
                messages.success(request, 'Цель обновлена.')
                return redirect('profile')

    history = Transaction.objects.filter(
        Q(from_user=user) | Q(to_user=user)
    ).select_related('from_user', 'to_user')[:20]

    return render(request, 'core/profile.html', {
        'form': goal_form,
        'edit_form': edit_form,
        'history': history,
    })


# ---------------------------------------------------------------------------
# Администратор / HR
# ---------------------------------------------------------------------------

@admin_required
def admin_dashboard(request):
    """Дашборд HR: вовлечённость, неактивные сотрудники, топ-5, новички, eNPS."""
    company = request.user.company
    today = timezone.now().date()

    employees = User.objects.filter(company=company)
    total_employees = employees.count()

    month_transactions = Transaction.objects.filter(
        type='give', from_user__company=company, year=today.year, month=today.month
    )

    active_user_ids = set(month_transactions.values_list('from_user_id', flat=True)) | \
        set(month_transactions.values_list('to_user_id', flat=True))
    active_count = len([uid for uid in active_user_ids if uid is not None])
    engagement_percent = round(active_count / total_employees * 100, 1) if total_employees else 0

    inactive_employees = employees.exclude(id__in=active_user_ids)

    top_receivers = month_transactions.values(
        'to_user__id', 'to_user__first_name', 'to_user__last_name', 'to_user__department'
    ).annotate(total=Sum('amount')).order_by('-total')[:5]

    top_senders = month_transactions.values(
        'from_user__id', 'from_user__first_name', 'from_user__last_name', 'from_user__department'
    ).annotate(total=Sum('amount')).order_by('-total')[:5]

    # Лучший новичок: стаж до 3 месяцев, больше всех получил монет за всё время
    three_months_ago = today.replace(day=1)
    # вычисляем дату "минус 3 месяца" простым способом
    month_index = today.month - 3
    year_offset = 0
    if month_index <= 0:
        month_index += 12
        year_offset = -1
    three_months_ago = date(today.year + year_offset, month_index, 1)

    newcomers = employees.filter(hire_date__gte=three_months_ago)
    best_newcomer = None
    if newcomers.exists():
        best_newcomer_row = Transaction.objects.filter(
            type='give', to_user__in=newcomers
        ).values(
            'to_user__id', 'to_user__first_name', 'to_user__last_name', 'to_user__department'
        ).annotate(total=Sum('amount')).order_by('-total').first()
        best_newcomer = best_newcomer_row

    latest_survey = ENPSSurvey.objects.filter(company=company).first()

    # --- Сплочённость отделов ---
    department_cohesion = _calculate_department_cohesion(company, today.year, today.month)

    # --- Уведомления о смене тарифа ---
    from .models import PlanChangeHistory
    unread_plan_changes = PlanChangeHistory.objects.filter(
        company=company, acknowledged=False
    ).order_by('-changed_at')

    if request.method == 'POST' and request.POST.get('action') == 'acknowledge_plan_change':
        PlanChangeHistory.objects.filter(company=company, acknowledged=False).update(acknowledged=True)
        return redirect('admin_dashboard')

    return render(request, 'core/admin_dashboard.html', {
        'total_employees': total_employees,
        'engagement_percent': engagement_percent,
        'active_count': active_count,
        'inactive_employees': inactive_employees,
        'top_receivers': top_receivers,
        'top_senders': top_senders,
        'best_newcomer': best_newcomer,
        'latest_survey': latest_survey,
        'department_cohesion': department_cohesion,
        'unread_plan_changes': unread_plan_changes,
    })


def _calculate_department_cohesion(company, year, month):
    """
    Считает сплочённость каждого отдела компании за указанный месяц.

    Сплочённость = % сотрудников отдела, которые подарили ИЛИ получили
    монетки от коллеги из того же отдела (внутриотдельское взаимодействие),
    округлённый до целого числа. Отделы без сотрудников не включаются.
    Возвращает список словарей, отсортированный по убыванию сплочённости.
    """
    departments = User.objects.filter(
        company=company
    ).exclude(department='').values_list('department', flat=True).distinct()

    result = []
    for dept in departments:
        dept_users = User.objects.filter(company=company, department=dept, is_active=True)
        dept_user_ids = set(dept_users.values_list('id', flat=True))
        total = len(dept_user_ids)
        if total == 0:
            continue

        # Транзакции "give" за месяц, где ОБЕ стороны — сотрудники этого отдела
        internal_transactions = Transaction.objects.filter(
            type='give', year=year, month=month,
            from_user_id__in=dept_user_ids, to_user_id__in=dept_user_ids,
        )

        engaged_ids = set(internal_transactions.values_list('from_user_id', flat=True)) | \
                      set(internal_transactions.values_list('to_user_id', flat=True))

        cohesion_percent = round(len(engaged_ids) / total * 100) if total else 0

        result.append({
            'department': dept,
            'total_employees': total,
            'engaged_employees': len(engaged_ids),
            'cohesion_percent': cohesion_percent,
        })

    result.sort(key=lambda x: x['cohesion_percent'], reverse=True)
    return result


@admin_required
def admin_employees(request):
    """Список сотрудников компании с балансами и отделами."""
    employees = User.objects.filter(company=request.user.company).order_by('last_name', 'first_name')
    return render(request, 'core/admin_employees.html', {'employees': employees})


@admin_required
def admin_rewards(request):
    """CRUD-список наград компании."""
    rewards = Reward.objects.filter(company=request.user.company)

    if request.method == 'POST':
        form = RewardForm(request.POST)
        if form.is_valid():
            reward = form.save(commit=False)
            reward.company = request.user.company
            reward.save()
            messages.success(request, f'Награда «{reward.name}» создана.')
            return redirect('admin_rewards')
    else:
        form = RewardForm()

    return render(request, 'core/admin_rewards.html', {'rewards': rewards, 'form': form})


@admin_required
def admin_reward_edit(request, pk):
    """Редактирование/деактивация награды."""
    reward = get_object_or_404(Reward, pk=pk, company=request.user.company)

    if request.method == 'POST':
        form = RewardForm(request.POST, instance=reward)
        if form.is_valid():
            form.save()
            messages.success(request, 'Награда обновлена.')
            return redirect('admin_rewards')
    else:
        form = RewardForm(instance=reward)

    return render(request, 'core/admin_reward_edit.html', {'form': form, 'reward': reward})


@admin_required
def admin_grant_coins(request, user_id):
    """Ручное начисление монет сотруднику администратором."""
    target = get_object_or_404(User, pk=user_id, company=request.user.company)

    if request.method == 'POST':
        try:
            amount = int(request.POST.get('amount', 0))
        except ValueError:
            amount = 0

        if amount > 0:
            today = timezone.now().date()
            with db_transaction.atomic():
                target.balance += amount
                target.save(update_fields=['balance'])
                Transaction.objects.create(
                    from_user=request.user,
                    to_user=target,
                    amount=amount,
                    comment='Начисление администратором',
                    type='admin',
                    month=today.month,
                    year=today.year,
                )
            messages.success(request, f'{target} получил {amount} монет.')
        else:
            messages.error(request, 'Укажите положительное количество монет.')

    return redirect('admin_employees')


@admin_required
@admin_required
def enps_start(request):
    """Запуск нового опроса eNPS — создаёт опрос и рассылает сотрудникам."""
    if request.method == 'POST':
        survey = ENPSSurvey.objects.create(company=request.user.company)

        # Рассылка: Telegram-уведомление всем сотрудникам компании
        try:
            from .telegram_bot import notify_enps_survey_started
            employees = User.objects.filter(
                company=request.user.company, is_active=True
            ).exclude(pk=request.user.pk)
            notified = 0
            for emp in employees:
                if emp.telegram_chat_id:
                    notify_enps_survey_started(emp, survey)
                    notified += 1
        except Exception:
            notified = 0

        messages.success(
            request,
            f'Опрос eNPS запущен и появился у сотрудников на главной странице.'
            + (f' Уведомлено в Telegram: {notified} чел.' if notified else '')
        )
        return redirect('enps_detail', pk=survey.pk)

    surveys = ENPSSurvey.objects.filter(company=request.user.company)
    return render(request, 'core/enps_start.html', {'surveys': surveys})


@admin_required
def enps_detail(request, pk):
    survey = get_object_or_404(ENPSSurvey, pk=pk, company=request.user.company)
    survey_link = 'https://www.24spasibo.ru' + reverse('enps_respond', args=[survey.pk])
    answered_count = survey.participations.count()
    total_employees = User.objects.filter(company=survey.company, is_active=True).count()
    return render(request, 'core/enps_detail.html', {
        'survey': survey,
        'survey_link': survey_link,
        'answered_count': answered_count,
        'total_employees': total_employees,
    })


@login_required
def enps_respond(request, pk):
    """
    Форма ответа на опрос eNPS.

    Доступна только авторизованным сотрудникам той же компании, что и опрос.
    Каждый сотрудник может ответить только один раз (ENPSParticipation).
    Сам ответ (балл, комментарий) анонимен — не связан с пользователем в данных.
    """
    from .models import ENPSParticipation

    survey = get_object_or_404(ENPSSurvey, pk=pk)

    if request.user.company_id != survey.company_id:
        messages.error(request, 'Этот опрос недоступен для вашей компании.')
        return redirect('dashboard')

    already_answered = ENPSParticipation.objects.filter(survey=survey, user=request.user).exists()
    submitted = False

    if request.method == 'POST' and not already_answered:
        form = ENPSResponseForm(request.POST)
        if form.is_valid():
            with db_transaction.atomic():
                survey.responses.append({
                    'score': form.cleaned_data['score'],
                    'comment': form.cleaned_data['comment'],
                })
                survey.recalculate_average()
                survey.save(update_fields=['responses', 'average_score'])

                ENPSParticipation.objects.create(survey=survey, user=request.user)

            submitted = True
            already_answered = True
            form = ENPSResponseForm()
    else:
        form = ENPSResponseForm()

    return render(request, 'core/enps_respond.html', {
        'survey': survey, 'form': form, 'submitted': submitted,
        'already_answered': already_answered,
    })


# ---------------------------------------------------------------------------
# Управление сотрудниками (добавление, редактирование, импорт)
# ---------------------------------------------------------------------------

@admin_required
def admin_employee_add(request):
    """Добавление сотрудников: один / несколько / импорт."""
    from .forms import EmployeeAddForm
    import csv, io

    company = request.user.company
    today   = timezone.now().date()
    form    = EmployeeAddForm(company=company)
    import_results = None

    if request.method == 'POST':
        method = request.POST.get('method', 'single')

        # --- Один сотрудник ---
        if method == 'single':
            form = EmployeeAddForm(request.POST, company=company)
            if form.is_valid():
                data = form.cleaned_data
                user = User.objects.create_user(
                    username=data['email'], email=data['email'],
                    password=data['password'],
                    first_name=data['first_name'],
                    last_name=data.get('last_name', ''),
                    company=company,
                    role=data.get('role', 'employee'),
                    department=data.get('department', ''),
                    hire_date=data.get('hire_date') or today,
                    balance=settings.MONTHLY_COIN_ALLOCATION,
                    last_monthly_allocation=today,
                )
                messages.success(request, f'Сотрудник {user.get_full_name() or user.email} добавлен.')
                return redirect('admin_employees')

        # --- Несколько сразу (bulk) ---
        elif method == 'bulk':
            first_names  = request.POST.getlist('bulk_first_name')
            last_names   = request.POST.getlist('bulk_last_name')
            emails       = request.POST.getlist('bulk_email')
            departments  = request.POST.getlist('bulk_department')
            password     = request.POST.get('bulk_password', 'Spasibo2024!').strip() or 'Spasibo2024!'
            role         = request.POST.get('bulk_role', 'employee')
            created      = 0

            for fn, ln, email, dept in zip(first_names, last_names, emails, departments):
                fn    = fn.strip()
                email = email.strip().lower()
                if not fn or not email or '@' not in email:
                    continue
                if User.objects.filter(email=email).exists():
                    continue
                User.objects.create_user(
                    username=email, email=email, password=password,
                    first_name=fn, last_name=ln.strip(),
                    company=company, role=role,
                    department=dept.strip(),
                    hire_date=today,
                    balance=settings.MONTHLY_COIN_ALLOCATION,
                    last_monthly_allocation=today,
                )
                created += 1

            if created:
                messages.success(request, f'Добавлено {created} сотрудников.')
                return redirect('admin_employees')
            else:
                messages.warning(request, 'Не добавлено ни одного сотрудника. Проверьте заполненность полей.')

        # --- Импорт из файла ---
        elif method == 'import':
            uploaded         = request.FILES.get('file')
            default_password = request.POST.get('default_password', 'Spasibo2024!').strip() or 'Spasibo2024!'
            created = 0; skipped = 0; errors = []

            if uploaded:
                filename = uploaded.name.lower()
                rows = []
                try:
                    if filename.endswith('.csv'):
                        content = uploaded.read().decode('utf-8-sig')
                        rows    = list(csv.DictReader(io.StringIO(content)))
                    elif filename.endswith(('.xlsx', '.xls')):
                        import openpyxl
                        wb      = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
                        ws      = wb.active
                        headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
                except Exception as e:
                    errors.append(f'Ошибка чтения файла: {e}')

                FIELD_MAP = {
                    'email': 'email',
                    'first_name': 'first_name', 'имя': 'first_name',
                    'last_name': 'last_name',   'фамилия': 'last_name',
                    'department': 'department',  'отдел': 'department',
                    'hire_date': 'hire_date',    'дата_найма': 'hire_date',
                    'role': 'role',              'роль': 'role',
                }
                for i, raw_row in enumerate(rows, start=2):
                    row   = {FIELD_MAP.get(k.strip().lower(), k): v for k, v in raw_row.items()}
                    email = row.get('email', '').strip().lower()
                    fn    = row.get('first_name', '').strip()
                    if not email or '@' not in email or not fn:
                        skipped += 1; continue
                    if User.objects.filter(email=email).exists():
                        skipped += 1; continue
                    hire_date = None
                    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
                        try:
                            from datetime import datetime as dt
                            hire_date = dt.strptime(row.get('hire_date',''), fmt).date()
                            break
                        except ValueError:
                            pass
                    role = row.get('role', 'employee').strip()
                    if role not in ('employee', 'admin'):
                        role = 'employee'
                    try:
                        User.objects.create_user(
                            username=email, email=email, password=default_password,
                            first_name=fn, last_name=row.get('last_name','').strip(),
                            company=company, role=role,
                            department=row.get('department','').strip(),
                            hire_date=hire_date or today,
                            balance=settings.MONTHLY_COIN_ALLOCATION,
                            last_monthly_allocation=today,
                        )
                        created += 1
                    except Exception as e:
                        errors.append(f'Строка {i} ({email}): {e}')

                import_results = {'created': created, 'skipped': skipped, 'errors': errors}
                if created:
                    messages.success(request, f'Импорт завершён: создано {created} сотрудников.')

    return render(request, 'core/admin_employee_add.html', {
        'form': form,
        'import_results': import_results,
    })


@admin_required
def admin_employee_edit(request, user_id):
    """Редактирование / деактивация сотрудника."""
    from .forms import EmployeeEditForm
    employee = get_object_or_404(User, pk=user_id, company=request.user.company)
    if request.method == 'POST':
        form = EmployeeEditForm(request.POST, instance=employee, company=request.user.company)
        if form.is_valid():
            emp = form.save(commit=False)
            emp.is_active = request.POST.get('is_active') == '1'
            new_pw = request.POST.get('new_password', '').strip()
            if new_pw:
                emp.set_password(new_pw)
            emp.save()
            messages.success(request, 'Данные сотрудника обновлены.')
            return redirect('admin_employees')
    else:
        form = EmployeeEditForm(instance=employee, company=request.user.company)
    return render(request, 'core/admin_employee_edit.html', {'form': form, 'employee': employee})


@admin_required
def admin_employee_import(request):
    """Импорт сотрудников из CSV или Excel."""
    import csv
    import io

    import_results = None

    if request.method == 'POST':
        uploaded = request.FILES.get('file')
        default_password = request.POST.get('default_password', 'Spasibo2024!').strip() or 'Spasibo2024!'

        if not uploaded:
            messages.error(request, 'Выберите файл для загрузки.')
        else:
            filename = uploaded.name.lower()
            created = 0
            skipped = 0
            errors = []

            # Определяем формат и читаем строки
            rows = []
            try:
                if filename.endswith('.csv'):
                    content = uploaded.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(content))
                    rows = list(reader)
                elif filename.endswith(('.xlsx', '.xls')):
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
                        ws = wb.active
                        headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
                    except ImportError:
                        errors.append('Для импорта Excel установите openpyxl: pip install openpyxl')
                else:
                    errors.append('Неподдерживаемый формат. Используйте .csv, .xlsx или .xls')
            except Exception as e:
                errors.append(f'Ошибка чтения файла: {e}')

            # Маппинг заголовков (рус/англ)
            FIELD_MAP = {
                'email': 'email',
                'first_name': 'first_name', 'имя': 'first_name', 'имя ': 'first_name',
                'last_name': 'last_name', 'фамилия': 'last_name',
                'department': 'department', 'отдел': 'department',
                'hire_date': 'hire_date', 'дата_найма': 'hire_date', 'дата найма': 'hire_date',
                'role': 'role', 'роль': 'role',
            }

            today = timezone.now().date()
            company = request.user.company

            for i, raw_row in enumerate(rows, start=2):
                row = {FIELD_MAP.get(k.strip().lower(), k.strip().lower()): v for k, v in raw_row.items()}
                email = row.get('email', '').strip().lower()
                first_name = row.get('first_name', '').strip()

                if not email or '@' not in email:
                    skipped += 1
                    continue
                if not first_name:
                    skipped += 1
                    errors.append(f'Строка {i}: нет имени для {email}')
                    continue
                if User.objects.filter(email=email).exists():
                    skipped += 1
                    continue

                # Парсим дату найма
                hire_date = None
                hire_raw = row.get('hire_date', '').strip()
                if hire_raw:
                    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
                        try:
                            from datetime import datetime
                            hire_date = datetime.strptime(hire_raw, fmt).date()
                            break
                        except ValueError:
                            pass

                role = row.get('role', 'employee').strip()
                if role not in ('employee', 'admin'):
                    role = 'employee'

                try:
                    User.objects.create_user(
                        username=email,
                        email=email,
                        password=default_password,
                        first_name=first_name,
                        last_name=row.get('last_name', '').strip(),
                        company=company,
                        role=role,
                        department=row.get('department', '').strip(),
                        hire_date=hire_date or today,
                        balance=settings.MONTHLY_COIN_ALLOCATION,
                        last_monthly_allocation=today,
                    )
                    created += 1
                except Exception as e:
                    errors.append(f'Строка {i} ({email}): {e}')

            import_results = {'created': created, 'skipped': skipped, 'errors': errors}
            if created:
                messages.success(request, f'Импорт завершён: создано {created} сотрудников.')

    return render(request, 'core/admin_employee_import.html', {
        'form': None,
        'import_results': import_results,
    })


@admin_required
def admin_employee_import_template(request):
    """Скачать шаблон CSV для импорта."""
    import csv
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="import_template.csv"'
    response.write('\ufeff')  # BOM for Excel
    writer = csv.writer(response)
    writer.writerow(['email', 'first_name', 'last_name', 'department', 'hire_date', 'role'])
    writer.writerow(['ivan@company.ru', 'Иван', 'Петров', 'Backend', '2024-01-15', 'employee'])
    writer.writerow(['anna@company.ru', 'Анна', 'Смирнова', 'Design', '2023-06-01', 'employee'])
    return response


# ---------------------------------------------------------------------------
# Онбординг новой компании (3 шага)
# ---------------------------------------------------------------------------

REWARD_TEMPLATES = [
    {'icon': '👕', 'name': 'Фирменный мерч', 'price': 25, 'category': 'material'},
    {'icon': '🍽️', 'name': 'Обед с руководителем', 'price': 100, 'category': 'event'},
    {'icon': '🌴', 'name': 'Дополнительный выходной', 'price': 40, 'category': 'wellbeing'},
    {'icon': '📚', 'name': 'Онлайн-курс на выбор', 'price': 60, 'category': 'development'},
    {'icon': '☕', 'name': 'Сертификат в кофейню', 'price': 15, 'category': 'material'},
    {'icon': '🎮', 'name': 'Ранний выход в пятницу', 'price': 20, 'category': 'wellbeing'},
]


@login_required
def onboarding(request):
    """Визард онбординга для нового HR-администратора (3 шага)."""
    user = request.user

    # Только для администраторов
    if user.role != 'admin':
        return redirect('dashboard')

    step = int(request.GET.get('step', 1))

    if request.method == 'POST':
        current_step = int(request.POST.get('step', 1))

        # --- Шаг 1: данные компании ---
        if current_step == 1:
            company = user.company
            if company:
                company_name = request.POST.get('company_name', '').strip()
                if company_name:
                    company.name = company_name
                    company.save(update_fields=['name'])

            user.first_name = request.POST.get('first_name', user.first_name).strip()
            user.last_name  = request.POST.get('last_name', user.last_name).strip()
            user.department = request.POST.get('department', user.department).strip()
            user.save(update_fields=['first_name', 'last_name', 'department'])

            return redirect(f"{request.path}?step=2")

        # --- Шаг 2: награды ---
        elif current_step == 2:
            company = user.company

            # Шаблонные награды
            for raw in request.POST.getlist('template_rewards'):
                parts = raw.split('|')
                if len(parts) == 3:
                    name, price_str, category = parts
                    try:
                        price = int(price_str)
                    except ValueError:
                        price = 25
                    Reward.objects.get_or_create(
                        company=company, name=name,
                        defaults={'price': price, 'category': category, 'is_active': True}
                    )

            # Своя награда
            custom_name  = request.POST.get('custom_name', '').strip()
            custom_price = request.POST.get('custom_price', '').strip()
            if custom_name and custom_price:
                try:
                    Reward.objects.get_or_create(
                        company=company, name=custom_name,
                        defaults={'price': int(custom_price), 'category': 'material', 'is_active': True}
                    )
                except ValueError:
                    pass

            return redirect(f"{request.path}?step=3")

        # --- Шаг 3: сотрудники ---
        elif current_step == 3:
            company  = user.company
            method   = request.POST.get('method', 'manual')
            today    = timezone.now().date()
            created  = 0

            if method == 'manual':
                first_names = request.POST.getlist('emp_first_name')
                last_names  = request.POST.getlist('emp_last_name')
                emails      = request.POST.getlist('emp_email')

                for fn, ln, email in zip(first_names, last_names, emails):
                    fn    = fn.strip()
                    email = email.strip().lower()
                    if not fn or not email or '@' not in email:
                        continue
                    if User.objects.filter(email=email).exists():
                        continue
                    User.objects.create_user(
                        username=email, email=email, password='Spasibo2024!',
                        first_name=fn, last_name=ln.strip(),
                        company=company, role='employee',
                        hire_date=today,
                        balance=settings.MONTHLY_COIN_ALLOCATION,
                        last_monthly_allocation=today,
                    )
                    created += 1

            elif method == 'import':
                import csv, io
                uploaded = request.FILES.get('file')
                if uploaded:
                    filename = uploaded.name.lower()
                    rows = []
                    try:
                        if filename.endswith('.csv'):
                            content = uploaded.read().decode('utf-8-sig')
                            reader  = csv.DictReader(io.StringIO(content))
                            rows    = list(reader)
                        elif filename.endswith(('.xlsx', '.xls')):
                            import openpyxl
                            wb      = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
                            ws      = wb.active
                            headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
                    except Exception:
                        pass

                    FIELD_MAP = {
                        'email': 'email',
                        'first_name': 'first_name', 'имя': 'first_name',
                        'last_name': 'last_name',   'фамилия': 'last_name',
                        'department': 'department',  'отдел': 'department',
                    }
                    for raw_row in rows:
                        row   = {FIELD_MAP.get(k.strip().lower(), k): v for k, v in raw_row.items()}
                        email = row.get('email', '').strip().lower()
                        fn    = row.get('first_name', '').strip()
                        if not email or '@' not in email or not fn:
                            continue
                        if User.objects.filter(email=email).exists():
                            continue
                        User.objects.create_user(
                            username=email, email=email, password='Spasibo2024!',
                            first_name=fn, last_name=row.get('last_name', '').strip(),
                            company=company, role='employee',
                            department=row.get('department', '').strip(),
                            hire_date=today,
                            balance=settings.MONTHLY_COIN_ALLOCATION,
                            last_monthly_allocation=today,
                        )
                        created += 1

            if created:
                messages.success(request, f'Готово! Добавлено {created} сотрудников. Платформа настроена 🎉')
            else:
                messages.success(request, 'Платформа настроена! Сотрудников можно добавить позже.')

            return redirect('dashboard')

    return render(request, 'core/onboarding.html', {
        'step': step,
        'reward_templates': REWARD_TEMPLATES,
    })


# ---------------------------------------------------------------------------
# Пригласительные ссылки
# ---------------------------------------------------------------------------

@admin_required
def admin_invites(request):
    """Управление пригласительными ссылками."""
    company = request.user.company

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create':
            days = int(request.POST.get('days_valid', 30))
            invite = CompanyInvite.generate(company, request.user, days_valid=days)

            # AJAX-запрос (например из онбординга) получает JSON с готовой ссылкой
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                link = request.build_absolute_uri(f'/join/{invite.token}/')
                return JsonResponse({'link': link, 'token': invite.token})

            messages.success(request, 'Пригласительная ссылка создана.')

        elif action == 'deactivate':
            invite_id = request.POST.get('invite_id')
            CompanyInvite.objects.filter(pk=invite_id, company=company).update(is_active=False)
            messages.success(request, 'Ссылка отозвана.')

        return redirect('admin_invites')

    invites = CompanyInvite.objects.filter(company=company)
    return render(request, 'core/admin_invites.html', {'invites': invites})


def invite_register(request, token):
    """Регистрация сотрудника по пригласительной ссылке."""

    try:
        invite = CompanyInvite.objects.select_related('company').get(token=token)
    except CompanyInvite.DoesNotExist:
        return render(request, 'core/invite_invalid.html', status=404)

    if not invite.is_valid:
        return render(request, 'core/invite_invalid.html', status=410)

    error     = None
    form_data = {}

    if request.method == 'POST':
        first_name  = request.POST.get('first_name', '').strip()
        last_name   = request.POST.get('last_name', '').strip()
        email       = request.POST.get('email', '').strip().lower()
        department  = request.POST.get('department', '').strip()
        password    = request.POST.get('password', '')
        password2   = request.POST.get('password2', '')

        form_data = {
            'first_name': first_name, 'last_name': last_name,
            'email': email, 'department': department,
        }

        # Валидация
        if not request.POST.get('consent'):
            error = 'Необходимо принять пользовательское соглашение и согласие на обработку данных.'
        elif not first_name:
            error = 'Введите имя.'
        elif not email or '@' not in email:
            error = 'Введите корректный email.'
        elif User.objects.filter(email=email).exists():
            error = 'Пользователь с таким email уже зарегистрирован. Попробуйте войти.'
        elif len(password) < 6:
            error = 'Пароль должен содержать минимум 6 символов.'
        elif password != password2:
            error = 'Пароли не совпадают.'

        if not error:
            today = timezone.now().date()
            with db_transaction.atomic():
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    company=invite.company,
                    role='employee',
                    department=department,
                    hire_date=today,
                    balance=settings.MONTHLY_COIN_ALLOCATION,
                    last_monthly_allocation=today,
                )
                invite.uses_count += 1
                invite.save(update_fields=['uses_count'])

            # Автоматический вход
            from django.contrib.auth import login as auth_login
            auth_login(request, user)
            messages.success(request, f'Добро пожаловать в «{invite.company.name}»! 🎉')
            return redirect('dashboard')

    return render(request, 'core/invite_register.html', {
        'invite': invite,
        'error': error,
        'form_data': form_data,
    })


# ---------------------------------------------------------------------------
# Настройка бонусов за стаж (администратор)
# ---------------------------------------------------------------------------

@admin_required
def admin_seniority_bonuses(request):
    """Управление правилами начисления бонусов за стаж."""
    from .models import SeniorityBonus
    company = request.user.company

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create':
            try:
                days   = int(request.POST.get('days_required', 0))
                amount = int(request.POST.get('coins_amount', 0))
            except ValueError:
                days = amount = 0

            if days > 0 and amount > 0:
                bonus, created = SeniorityBonus.objects.get_or_create(
                    company=company, days_required=days,
                    defaults={'coins_amount': amount},
                )
                if not created:
                    bonus.coins_amount = amount
                    bonus.is_active = True
                    bonus.save(update_fields=['coins_amount', 'is_active'])
                messages.success(request, f'Правило «{days} дней → {amount} монет» сохранено.')
            else:
                messages.error(request, 'Укажите положительные значения для дней и монет.')

        elif action == 'toggle':
            bonus_id = request.POST.get('bonus_id')
            bonus = SeniorityBonus.objects.filter(pk=bonus_id, company=company).first()
            if bonus:
                bonus.is_active = not bonus.is_active
                bonus.save(update_fields=['is_active'])

        elif action == 'delete':
            bonus_id = request.POST.get('bonus_id')
            SeniorityBonus.objects.filter(pk=bonus_id, company=company).delete()
            messages.success(request, 'Правило удалено.')

        return redirect('admin_seniority_bonuses')

    bonuses = SeniorityBonus.objects.filter(company=company)
    return render(request, 'core/admin_seniority_bonuses.html', {'bonuses': bonuses})


# ---------------------------------------------------------------------------
# Исполнение заявок на награды
# ---------------------------------------------------------------------------

@admin_required
def admin_reward_orders(request):
    """Список заявок на награды (покупок) с возможностью отметить исполнение."""
    company = request.user.company

    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        order = Transaction.objects.filter(
            pk=order_id, type='reward', to_user__isnull=True
        ).filter(from_user__company=company).first()

        if order and order.status == 'pending':
            order.status = 'fulfilled'
            order.fulfilled_at = timezone.now()
            order.fulfilled_by = request.user
            order.save(update_fields=['status', 'fulfilled_at', 'fulfilled_by'])
            messages.success(request, f'Награда «{order.reward.name if order.reward else order.comment}» отмечена как исполненная.')

            # Уведомление сотруднику
            try:
                from .telegram_bot import notify_reward_fulfilled
                if order.from_user and order.from_user.telegram_chat_id:
                    notify_reward_fulfilled(order.from_user, order)
            except Exception:
                pass

        return redirect('admin_reward_orders')

    pending_orders = Transaction.objects.filter(
        type='reward', status='pending', from_user__company=company
    ).select_related('from_user', 'reward').order_by('date')

    fulfilled_orders = Transaction.objects.filter(
        type='reward', status='fulfilled', from_user__company=company
    ).select_related('from_user', 'reward', 'fulfilled_by').order_by('-fulfilled_at')[:30]

    return render(request, 'core/admin_reward_orders.html', {
        'pending_orders': pending_orders,
        'fulfilled_orders': fulfilled_orders,
    })


# ---------------------------------------------------------------------------
# Обращение в поддержку
# ---------------------------------------------------------------------------

@login_required
def support_request(request):
    """Форма обращения в службу поддержки платформы."""
    from .models import SupportRequest

    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()

        if not subject or not message:
            messages.error(request, 'Заполните тему и сообщение.')
        else:
            req = SupportRequest.objects.create(
                user=request.user,
                company=request.user.company,
                subject=subject,
                message=message,
            )

            # Уведомление владельцу платформы через Telegram
            try:
                from .telegram_bot import send_message as tg_send
                owner_chat_id = settings.SUPPORT_TELEGRAM_CHAT_ID
                if owner_chat_id:
                    text = (
                        f'📨 <b>Новое обращение в поддержку #{req.pk}</b>\n\n'
                        f'От: <b>{request.user.get_full_name() or request.user.email}</b>\n'
                        f'Email: {request.user.email}\n'
                        f'Компания: {request.user.company.name if request.user.company else "—"}\n\n'
                        f'Тема: <b>{subject}</b>\n\n'
                        f'{message}'
                    )
                    tg_send(int(owner_chat_id), text)
            except Exception:
                pass  # не ронять страницу при ошибке Telegram

            messages.success(
                request,
                'Спасибо! Ваше обращение принято, мы ответим в ближайшее время.'
            )
            return redirect('support_request')

    my_requests = SupportRequest.objects.filter(user=request.user)[:10]
    return render(request, 'core/support_request.html', {'my_requests': my_requests})


# ---------------------------------------------------------------------------
# Юридические страницы
# ---------------------------------------------------------------------------

def privacy(request):
    return render(request, 'core/privacy.html')


def terms(request):
    return render(request, 'core/terms.html')
